import datetime
import csv
import openpyxl
import os
from function import CommonFunction as CommonFunction
from openpyxl.styles import Protection
from openpyxl.styles.borders import Border, Side, BORDER_THIN
from openpyxl.worksheet.filters import (FilterColumn, Filters)
from sharepoint import UploadFile
from variables import area, assetNbrLength, excelPath, ExcelLib, ExcelTargetLib, FilePassword, Font, TablesLib, TablesTargetLib, pathResultExcel, pathResultSAP, pathResultSummary
from variables import LeaseReconSchedulerPath, LeaseReconTemplatePath, sharepointFolderDetail

class Excel:
    def convertToXLSX(input_file, output_file):
        try:
            wb = openpyxl.Workbook()
            ws = wb.worksheets[0]

            with open(input_file,'r') as data:
                reader = csv.reader(data, delimiter='\t')
                for row in reader:
                    ws.append(row)

            wb.save(output_file)
        except Exception as errConvert:
            error = str(errConvert)
            CommonFunction.WriteLog(f"Failed convert to XLSX error with {error}")

    def CopyData(inSource, inSourceStart, inTargetSheet, inStartTargetCell, inStatus, inStatusColumn, inCompanyCode, inCompanyCodeColumn, Month, Year):
        try:
            CommonFunction.WriteLog(f"Open Source File: {inSource}")
            # For Depr Simulation read with header
            if inTargetSheet == 9 or inTargetSheet == 10:
                Header = True
            else:
                Header = False
                
            ExcelLib.open_workbook(path=inSource)
            Worksheet = ExcelLib.read_worksheet_as_table(start=inSourceStart, header= Header)
            ExcelLib.close_workbook()
            
            # Filter by Company Code
            if inCompanyCode != "" and inCompanyCodeColumn != "":
                TablesLib.filter_table_by_column(Worksheet, inCompanyCodeColumn, "==", inCompanyCode)
                CommonFunction.WriteLog(f"filtered company. Company: {inCompanyCode} - Column: {inCompanyCodeColumn}")

            # Filter by Status
            if inStatus != "" and inStatusColumn != "":
                TablesLib.filter_table_by_column(Worksheet, inStatusColumn, "==", inStatus)
                CommonFunction.WriteLog(f"filtered status. Status: {inStatus} - Column: {inStatusColumn}")
            
            #remove total in FBL3N
            if inTargetSheet == 7 or inTargetSheet == 8:
                TablesLib.filter_table_by_column(Worksheet, "H", "Not In", [None, " ", ""])
            # Write to Lease Month End Reconciliation
            ExcelTargetLib.set_active_worksheet(inTargetSheet)
            ExcelTargetLib.delete_rows(int(inStartTargetCell[1:len(inStartTargetCell)]),1048576)
                
            if inTargetSheet == 9 or inTargetSheet == 10:
                # Copy Depr Simulation
                # Month = "01"
                # Year = "2023"
                ColumnValue = "Depr. " + Month + "/" + Year
                i = len(Worksheet.columns)
                while i >= 0:
                    i = i - 1
                    if Worksheet.columns[i] != "Object" and Worksheet.columns[i] != ColumnValue:
                        TablesLib.pop_table_column(table=Worksheet, column=Worksheet.columns[i])
                TablesLib.filter_table_by_column(Worksheet, inStatusColumn, "!=", inStatus)
                ExcelTargetLib.set_cell_value(row=1, column="C", value=ColumnValue)
            
            ExcelTargetLib.set_cell_values(inStartTargetCell, Worksheet)
        except Exception as errCopy:
            error = str(errCopy)
            CommonFunction.WriteLog(f"Failed copy data with error: {error}")
        
    def FindStartRowNum(inSource, inHeaderValue, inColumn):
        try:
            ExcelLib.open_workbook(path=inSource)
            Worksheet = ExcelLib.read_worksheet_as_table()
            ExcelLib.close_workbook()
            StartRowNum = 1
            for data in Worksheet:
                if data[inColumn] == inHeaderValue:
                    StartRowNum = StartRowNum + 1
                    break
                StartRowNum = StartRowNum + 1
            return StartRowNum
        except Exception as errRowNum:
            error = str(errRowNum)
            CommonFunction.WriteLog(f"Failed find start row num with error: {error}")

    def ProcessingAssetTransactionSAP(month):
        try:
            TargetWorkSheet = ExcelTargetLib.read_worksheet_as_table(name="Asset Transaction SAP", start=1)
            # Filter Asset Column
            ExcelTargetLib.set_cell_value(row=7,column="W",value="Asset")
            ExcelTargetLib.set_cell_value(row=7,column="X",value="AsstValDat")
            StartRow = 8
            for i in range(len(TargetWorkSheet)):
                if str(TargetWorkSheet[i][1])[0:1] == "9":
                    AsstValDat = TargetWorkSheet[i+1][8]
                    if AsstValDat == None:
                        AsstValDat = TargetWorkSheet[i+2][8]
                        
                    if month != AsstValDat[3:5]:
                        ExcelTargetLib.set_cell_value(row=StartRow, column="W", value=int(TargetWorkSheet[i][1]))
                        ExcelTargetLib.set_cell_value(row=StartRow, column="X", value=AsstValDat)
                        StartRow = StartRow + 1
        except Exception as errAssetSAP:
            error = str(errAssetSAP)
            CommonFunction.WriteLog(f"Failed find start row num with error: {error}")

    def ProcessingCompletenesscheckUSGAAP(code,LastDate):
        CommonFunction.WriteLog(f"Processing Completeness check USGAAP {code} - {LastDate}")
        try:
            TargetWorkSheet = ExcelTargetLib.read_worksheet_as_table(name="AG SLAN", start=1)
            TablesTargetLib.filter_table_by_column(table=TargetWorkSheet, column="B",operator="==", value="active")
            ContractIDList = TablesTargetLib.get_table_column(table=TargetWorkSheet,column="D")
            
            ExcelTargetLib.set_active_worksheet(value="Completeness check US GAAP")
            ExcelTargetLib.set_cell_value(row=1,column="B",value=str(code))
            ExcelTargetLib.set_styles(range_string="B1", align_horizontal="right", bold=True, font_name=Font, size=8)
            ExcelTargetLib.set_cell_value(row=2,column="B",value=datetime.datetime.strptime(LastDate, '%d-%m-%Y').date(),name="Completeness check US GAAP",fmt="d-mmm-yy")
            StartRow = 7
            if area == "EU":
                fontSize = 11
                fontName= "Calibri"
            elif area == "LA":
                fontSize = 8
                fontName = Font
            elif area == "AP":
                fontSize = 8
                fontName = Font
            else:
                fontSize = 11
                fontName= "Calibri"
                
            if len(ContractIDList) > 0:
                #Column A Contract No.
                i = StartRow
                for ContractID in ContractIDList:
                    ExcelTargetLib.set_cell_value(row=i, column="A", value=ContractID, name="Completeness check US GAAP")
                    i = i + 1
                
                #Column B Asset Number
                ExcelTargetLib.set_cell_formula(range_string=f"B{StartRow}:B{len(ContractIDList)+StartRow-1}", formula="=VALUE(INDEX('Unit SLAN'!E:E,MATCH('Completeness check US GAAP'!A:A,'Unit SLAN'!H:H,0)))", transpose= True)
                
                #Column C Backdated contract creation?
                ExcelTargetLib.set_cell_formula(range_string=f"C{StartRow}:C{len(ContractIDList)+StartRow-1}", formula=f"=IFERROR(IF(VALUE(VLOOKUP(VALUE(B{StartRow}),'Asset Transaction SAP'!W:W,1,0)),\"yes\",\"no\"),\"no\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"C{StartRow}:C{len(ContractIDList)+StartRow-1}", align_horizontal="center")

                #Column D Remaining Lease Contract (Days)
                ExcelTargetLib.set_cell_formula(range_string=f"D{StartRow}:D{len(ContractIDList)+StartRow-1}", formula=f"=IF(ISERROR(E{StartRow}-$B$2),\"Error in SLAN\",IF((E{StartRow}-$B$2)<0,IF(AND(YEAR(E{StartRow})=YEAR($B$2),MONTH(E{StartRow})=MONTH($B$2)),\"Contract end in current month\",(E{StartRow}-$B$2)),(E{StartRow}-$B$2)))", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"D{StartRow}:D{len(ContractIDList)+StartRow-1}", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')
                
                #Column E Contract End Date
                ExcelTargetLib.set_cell_formula(range_string=f"E{StartRow}:E{len(ContractIDList)+StartRow-1}", formula=f"=IFERROR(VLOOKUP(A{StartRow},'Liability SLAN US GAAP'!A:P,10,0),\"Inactive in SLAN\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"E{StartRow}:E{len(ContractIDList)+StartRow-1}", align_horizontal="center", number_format="m/d/yyyy")

                #Column F Depreciation
                ExcelTargetLib.set_cell_formula(range_string=f"F{StartRow}:F{len(ContractIDList)+StartRow-1}", formula=f"=IFERROR(IF(VLOOKUP(B{StartRow},'Depr Simulation SAP US GAAP'!A:C,3,0)<0,\"Posted\",\"Not Posted\"),\"Already Retired\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"F{StartRow}:F{len(ContractIDList)+StartRow-1}", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')

                #Column G Check Double Posting
                ExcelTargetLib.set_cell_formula(range_string=f"G{StartRow}:G{len(ContractIDList)+StartRow-1}", formula=f"=COUNTIFS('FBL3N SAP US GAAP'!B:B,20302500,'FBL3N SAP US GAAP'!G:G,A{StartRow},'FBL3N SAP US GAAP'!F:F,\"Accrual\",'FBL3N SAP US GAAP'!E:E,\"IS\")+COUNTIFS('FBL3N SAP US GAAP'!B:B,20312500,'FBL3N SAP US GAAP'!G:G,A{StartRow},'FBL3N SAP US GAAP'!F:F,\"Accrual\",'FBL3N SAP US GAAP'!E:E,\"IS\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"G{StartRow}:G{len(ContractIDList)+StartRow-1}", cell_fill="#FFF2CB", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')
                
                #Column H Liatbility balance(ST/LT)
                ExcelTargetLib.set_cell_formula(range_string=f"H{StartRow}:H{len(ContractIDList)+StartRow-1}", formula=f"=IF(ROUND(SUMIF('Liability SLAN US GAAP'!A:A,'Completeness check US GAAP'!A{StartRow},'Liability SLAN US GAAP'!U:U)+SUMIF('Liability SLAN US GAAP'!A:A,'Completeness check US GAAP'!A{StartRow},'Liability SLAN US GAAP'!V:V),0)>0,\"Yes\",\"No\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"H{StartRow}:H{len(ContractIDList)+StartRow-1}", align_horizontal="center", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')

                #Column I Check
                ExcelTargetLib.set_cell_formula(range_string=f"I{StartRow}:I{len(ContractIDList)+StartRow-1}", formula=f"=IF(D{StartRow}=\"Error in SLAN\",\"Check\",IF(AND(G{StartRow}=1,D{StartRow}>=0),\"Ok\",IF(AND(G{StartRow}=0,D{StartRow}>=0,H{StartRow}=\"No\"),\"Ok, No Future Payment\",\"Check\")))", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"I{StartRow}:I{len(ContractIDList)+StartRow-1}", align_horizontal="center", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')

                #Column J Check Double Posting
                ExcelTargetLib.set_cell_formula(range_string=f"J{StartRow}:J{len(ContractIDList)+StartRow-1}", formula="=COUNTIFS('FBL3N SAP US GAAP'!B:B,20340200,'FBL3N SAP US GAAP'!G:G,'Completeness check US GAAP'!A:A,'FBL3N SAP US GAAP'!F:F,\"Payment\",'FBL3N SAP US GAAP'!E:E,\"IS\")+COUNTIFS('FBL3N SAP US GAAP'!B:B,20340400,'FBL3N SAP US GAAP'!G:G,'Completeness check US GAAP'!A:A,'FBL3N SAP US GAAP'!F:F,\"Payment\",'FBL3N SAP US GAAP'!E:E,\"IS\")+COUNTIFS('FBL3N SAP US GAAP'!B:B,20330100,'FBL3N SAP US GAAP'!G:G,'Completeness check US GAAP'!A:A,'FBL3N SAP US GAAP'!F:F,\"Payment\",'FBL3N SAP US GAAP'!E:E,\"IS\")+COUNTIFS('FBL3N SAP US GAAP'!B:B,20330200,'FBL3N SAP US GAAP'!G:G,'Completeness check US GAAP'!A:A,'FBL3N SAP US GAAP'!F:F,\"Payment\",'FBL3N SAP US GAAP'!E:E,\"IS\")+COUNTIFS('FBL3N SAP US GAAP'!B:B,20330300,'FBL3N SAP US GAAP'!G:G,'Completeness check US GAAP'!A:A,'FBL3N SAP US GAAP'!F:F,\"Payment\",'FBL3N SAP US GAAP'!E:E,\"IS\")+COUNTIFS('FBL3N SAP US GAAP'!B:B,20330400,'FBL3N SAP US GAAP'!G:G,'Completeness check US GAAP'!A:A,'FBL3N SAP US GAAP'!F:F,\"Payment\",'FBL3N SAP US GAAP'!E:E,\"IS\")+COUNTIFS('FBL3N SAP US GAAP'!B:B,20330600,'FBL3N SAP US GAAP'!G:G,'Completeness check US GAAP'!A:A,'FBL3N SAP US GAAP'!F:F,\"Payment\",'FBL3N SAP US GAAP'!E:E,\"IS\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"J{StartRow}:J{len(ContractIDList)+StartRow-1}", cell_fill="#FFF2CB", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')
                
                #Column K Payment Current Month?
                ExcelTargetLib.set_cell_formula(range_string=f"K{StartRow}:K{len(ContractIDList)+StartRow-1}", formula=f"=IF(IFERROR(VLOOKUP(A{StartRow},'Cash flow SLAN US GAAP'!A:U,21,0),0)<>0,\"Yes\",\"No\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"K{StartRow}:K{len(ContractIDList)+StartRow-1}", align_horizontal="center", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')

                #Column L Check
                ExcelTargetLib.set_cell_formula(range_string=f"L{StartRow}:L{len(ContractIDList)+StartRow-1}", formula=f"=IF(D{StartRow}=\"Error in SLAN\",\"Check\",IF(AND(J{StartRow}=1,D{StartRow}>=0,K{StartRow}=\"Yes\"),\"Ok\",IF(AND(J{StartRow}=0,D{StartRow}>=0,K{StartRow}=\"No\"),\"Ok, No Payment\",\"Check\")))", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"L{StartRow}:L{len(ContractIDList)+StartRow-1}", align_horizontal="center", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')

                #Column M Checking All
                ExcelTargetLib.set_cell_formula(range_string=f"M{StartRow}:M{len(ContractIDList)+StartRow-1}", formula=f"=IF(C{StartRow}=\"yes\",\"Check\",IF(AND(OR(I{StartRow}=\"Ok\",I{StartRow}=\"Ok, No Future Payment\"),OR(L{StartRow}=\"Ok\",L{StartRow}=\"Ok, No Payment\"),F{StartRow}=\"Posted\"),\"Ok\",\"Check\"))", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"M{StartRow}:M{len(ContractIDList)+StartRow-1}", align_horizontal="center", bold=True, number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')

                #Column N Remarks
                ExcelTargetLib.set_cell_formula(range_string=f"N{StartRow}:N{len(ContractIDList)+StartRow-1}", formula=f"=IF(M{StartRow}=\"Ok\",\"\",\"please provide comment\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"N{StartRow}:N{len(ContractIDList)+StartRow-1}", font_name=fontName, size=fontSize, color="#0563C1", underline="single", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')

                ExcelTargetLib.set_styles(range_string=f"A{StartRow}:M{len(ContractIDList)+StartRow-1}",font_name=Font,size=8)
                # ExcelTargetLib.delete_rows(start=7,end=7)
                ExcelTargetLib.auto_size_columns(start_column="N")
        except Exception as errCompletenessUSGAAP:
            error = str(errCompletenessUSGAAP)
            CommonFunction.WriteLog(f"Failed Processing Completeness US GAAP error with {error}")

    def ProcessingCompletenesscheckIFRS(code, LastDate):
        CommonFunction.WriteLog(f"Processing Completeness check IFRS {code} - {LastDate}")
        try:
            TargetWorkSheet = ExcelTargetLib.read_worksheet_as_table(name="AG SLAN", start=1)
            WorkingTable = TablesTargetLib.create_table(data=TargetWorkSheet, trim=True)
            TablesTargetLib.filter_table_by_column(table=WorkingTable, column="B",operator="==", value="active")
            ContractIDList = TablesTargetLib.get_table_column(table=WorkingTable,column="D")
            
            ExcelTargetLib.set_active_worksheet(value="Completeness check IFRS")
            ExcelTargetLib.set_cell_value(row=1, column="B", value=str(code))
            ExcelTargetLib.set_styles(range_string="B1", align_horizontal="right", bold=True, font_name=Font, size=8)
            ExcelTargetLib.set_cell_value(row=2, column="B", value=datetime.datetime.strptime(LastDate, '%d-%m-%Y').date(), name="Completeness check IFRS", fmt="d-mmm-yy")
            StartRow = 7
            
            if area == "EU":
                fontSize = 11
                fontName= "Calibri"
            elif area == "LA":
                fontSize = 8
                fontName = Font
            elif area == "AP":
                fontSize = 8
                fontName = Font
            else:
                fontSize = 11
                fontName= "Calibri"
                
            if len(ContractIDList) > 0:
            
                i = StartRow
                for ContractID in ContractIDList:
                    ExcelTargetLib.set_cell_value(row=i, column="A", value=ContractID, name="Completeness check IFRS")
                    i = i + 1

                ExcelTargetLib.set_cell_formula(range_string=f"B{StartRow}:B{len(ContractIDList)+StartRow-1}", formula="=VALUE(INDEX('Unit SLAN'!E:E,MATCH(A:A,'Unit SLAN'!H:H,0)))", transpose= True)

                ExcelTargetLib.set_cell_formula(range_string=f"C{StartRow}:C{len(ContractIDList)+StartRow-1}", formula=f"=IFERROR(IF(VALUE(VLOOKUP(VALUE(B{StartRow}),'Asset Transaction SAP'!W:W,1,0)),\"yes\",\"no\"),\"no\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"C{StartRow}:C{len(ContractIDList)+StartRow-1}", align_horizontal="center")

                ExcelTargetLib.set_cell_formula(range_string=f"D{StartRow}:D{len(ContractIDList)+StartRow-1}", formula=f"=IF(ISERROR(E{StartRow}-$B$2),\"Error in SLAN\",IF((E{StartRow}-$B$2)<0,IF(AND(YEAR(E{StartRow})=YEAR($B$2),MONTH(E{StartRow})=MONTH($B$2)),\"Contract end in current month\",(E{StartRow}-$B$2)),(E{StartRow}-$B$2)))", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"D{StartRow}:D{len(ContractIDList)+StartRow-1}", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')
                
                ExcelTargetLib.set_cell_formula(range_string=f"E{StartRow}:E{len(ContractIDList)+StartRow-1}", formula=f"=IFERROR(VLOOKUP(A{StartRow},'Liability SLAN IFRS'!A:P,10,0),\"Inactive in SLAN\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"E{StartRow}:E{len(ContractIDList)+StartRow-1}", align_horizontal="center", number_format="m/d/yyyy")

                ExcelTargetLib.set_cell_formula(range_string=f"F{StartRow}:F{len(ContractIDList)+StartRow-1}", formula=f"=IFERROR(IF(VLOOKUP(B{StartRow},'Depr Simulation SAP IFRS'!A:C,3,0)<0,\"Posted\",\"Not Posted\"),\"Already Retired\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"F{StartRow}:F{len(ContractIDList)+StartRow-1}", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')
                
                ExcelTargetLib.set_cell_formula(range_string=f"G{StartRow}:G{len(ContractIDList)+StartRow-1}", formula=f"=COUNTIFS('FBL3N SAP IFRS'!B:B,92030300,'FBL3N SAP IFRS'!G:G,A{StartRow},'FBL3N SAP IFRS'!F:F,\"Accrual\",'FBL3N SAP IFRS'!E:E,\"IS\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"G{StartRow}:G{len(ContractIDList)+StartRow-1}", cell_fill="#FFF2CB", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')
                
                ExcelTargetLib.set_cell_formula(range_string=f"H{StartRow}:H{len(ContractIDList)+StartRow-1}", formula=f"=IF(ROUND(SUMIF('Liability SLAN IFRS'!A:A,'Completeness check IFRS'!A{StartRow},'Liability SLAN IFRS'!U:U)+SUMIF('Liability SLAN IFRS'!A:A,'Completeness check IFRS'!A{StartRow},'Liability SLAN IFRS'!V:V),0)>0,\"Yes\",\"No\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"H{StartRow}:H{len(ContractIDList)+StartRow-1}", align_horizontal="center", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')

                ExcelTargetLib.set_cell_formula(range_string=f"I{StartRow}:I{len(ContractIDList)+StartRow-1}", formula=f"=IF(D{StartRow}=\"Error in SLAN\",\"Check\",IF(AND(G{StartRow}=1,D{StartRow}>=0),\"Ok\",IF(AND(G{StartRow}=0,D{StartRow}>=0,H{StartRow}=\"No\"),\"Ok, No Future Payment\",\"Check\")))", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"I{StartRow}:I{len(ContractIDList)+StartRow-1}", align_horizontal="center",number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')

                ExcelTargetLib.set_cell_formula(range_string=f"J{StartRow}:J{len(ContractIDList)+StartRow-1}", formula=f"=COUNTIFS('FBL3N SAP IFRS'!B:B,92034200,'FBL3N SAP IFRS'!G:G,'Completeness check IFRS'!A{StartRow},'FBL3N SAP IFRS'!F:F,\"Payment\",'FBL3N SAP IFRS'!E:E,\"IS\")+COUNTIFS('FBL3N SAP IFRS'!B:B,92034400,'FBL3N SAP IFRS'!G:G,'Completeness check IFRS'!A{StartRow},'FBL3N SAP IFRS'!F:F,\"Payment\",'FBL3N SAP IFRS'!E:E,\"IS\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"J{StartRow}:J{len(ContractIDList)+StartRow-1}", cell_fill="#FFF2CB", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')
                
                ExcelTargetLib.set_cell_formula(range_string=f"K{StartRow}:K{len(ContractIDList)+StartRow-1}", formula=f"=IF(IFERROR(VLOOKUP(A{StartRow},'Cash flow SLAN IFRS'!A:V,19,0)+VLOOKUP(A{StartRow},'Cash flow SLAN IFRS'!A:V,20,0),0)<>0,\"Yes\",\"No\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"K{StartRow}:K{len(ContractIDList)+StartRow-1}", align_horizontal="center", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')

                ExcelTargetLib.set_cell_formula(range_string=f"L{StartRow}:L{len(ContractIDList)+StartRow-1}", formula=f"=IF(D{StartRow}=\"Error in SLAN\",\"Check\",IF(AND(J{StartRow}=1,D{StartRow}>=0,K{StartRow}=\"Yes\"),\"Ok\",IF(AND(J{StartRow}=0,D{StartRow}>=0,K{StartRow}=\"No\"),\"Ok, No Payment\",\"Check\")))", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"L{StartRow}:L{len(ContractIDList)+StartRow-1}", align_horizontal="center", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')

                ExcelTargetLib.set_cell_formula(range_string=f"M{StartRow}:M{len(ContractIDList)+StartRow-1}", formula=f"=IF(C{StartRow}=\"yes\",\"Check\",IF(AND(OR(I{StartRow}=\"Ok\",I{StartRow}=\"Ok, No Future Payment\"),OR(L{StartRow}=\"Ok\",L{StartRow}=\"Ok, No Payment\"),F{StartRow}=\"Posted\"),\"Ok\",\"Check\"))", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"M{StartRow}:M{len(ContractIDList)+StartRow-1}", align_horizontal="center", bold=True, number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')

                ExcelTargetLib.set_cell_formula(range_string=f"N{StartRow}:N{len(ContractIDList)+StartRow-1}", formula=f"=IF(M{StartRow}=\"Ok\",\"\",\"please provide comment\")", transpose= True)
                ExcelTargetLib.set_styles(range_string=f"N{StartRow}:N{len(ContractIDList)+StartRow-1}", font_name=fontName, size=fontSize, color="#0563C1", underline="single", number_format='_(* #,##0_);_(* \(#,##0\);_(* "-"??_);_(@_)')

                ExcelTargetLib.set_styles(range_string=f"A{StartRow}:M{len(ContractIDList)+StartRow-1}",font_name=Font,size=8)
                # ExcelTargetLib.delete_rows(start=7,end=7)
                ExcelTargetLib.auto_size_columns(start_column="N")
        except Exception as errCompletenessIFRS:
            error = str(errCompletenessIFRS)
            CommonFunction.WriteLog(f"Failed Processing Completeness IFRS error with {error}")

    def HideSheet(inWorkBook, inWorkSheet):
        try:
            inWorkBook.move_sheet(sheet=inWorkSheet, offset=-1)
            WorkSheet = inWorkBook[inWorkSheet]
            WorkSheet.sheet_state = 'hidden'
        except Exception as errProtect:
            error = str(errProtect)
            CommonFunction.WriteLog(f"Failed hide sheet error with {error}")
            
    def ProtectSheet(inWorkBook, inWorkSheet, inPassword, inContractIDList):
        try:
            WorkSheet = inWorkBook[inWorkSheet]
            if inWorkSheet == "Completeness check US GAAP" or inWorkSheet == "Completeness check IFRS":
                z = 7
                for data in inContractIDList:
                    Cell = WorkSheet[f"N{z}"]
                    Cell.protection = Protection(locked=False)
                    z = z + 1

            if inWorkSheet == "AG SLAN":
                ColFilter = FilterColumn(colId=1)
                ColFilter.filters = Filters(filter=["active"])
                WorkSheet.auto_filter.filterColumn.append(ColFilter)
                for i, row in enumerate(WorkSheet.rows):
                    if i > 0:
                        Cell = WorkSheet[f"B{i+1}"]
                        if Cell.value != "active":
                            WorkSheet.row_dimensions[i+1].hidden = True

            WorkSheet.protection.autoFilter = False
            WorkSheet.protection.sort = False
            WorkSheet.protection.sheet = True
            WorkSheet.protection.password = inPassword
        except Exception as errProtect:
            error = str(errProtect)
            CommonFunction.WriteLog(f"Failed protect sheet error with {error}")

    def ProtectFiles(inWorkBook, inPassword, inContractIDList):
        try:
            CommonFunction.WriteLog(f"Start Lock Sheets")
            # Lock Sheets
            Excel.ProtectSheet(inWorkBook, "AG SLAN", inPassword, inContractIDList)
            Excel.ProtectSheet(inWorkBook, "Unit SLAN", inPassword, inContractIDList)
            Excel.ProtectSheet(inWorkBook, "Cash flow SLAN US GAAP", inPassword, inContractIDList)
            Excel.ProtectSheet(inWorkBook, "Cash flow SLAN IFRS", inPassword, inContractIDList)
            Excel.ProtectSheet(inWorkBook, "Liability SLAN US GAAP", inPassword, inContractIDList)
            Excel.ProtectSheet(inWorkBook, "Liability SLAN IFRS", inPassword, inContractIDList)
            Excel.ProtectSheet(inWorkBook, "FBL3N SAP US GAAP", inPassword, inContractIDList)
            Excel.ProtectSheet(inWorkBook, "FBL3N SAP IFRS", inPassword, inContractIDList)
            Excel.ProtectSheet(inWorkBook, "Depr Simulation SAP US GAAP", inPassword, inContractIDList)
            Excel.ProtectSheet(inWorkBook, "Depr Simulation SAP IFRS", inPassword, inContractIDList)
            Excel.ProtectSheet(inWorkBook, "Asset Transaction SAP", inPassword, inContractIDList)
            Excel.ProtectSheet(inWorkBook, "Completeness check US GAAP", inPassword, inContractIDList)
            Excel.ProtectSheet(inWorkBook, "Completeness check IFRS", inPassword, inContractIDList)
            CommonFunction.WriteLog(f"Finish Lock Sheets")
        except Exception as errProtect:
            error = str(errProtect)
            CommonFunction.WriteLog(f"Failed protect files error with {error}")
        
    def setCellFormat(type, startRow):
        try:
            CommonFunction.WriteLog(f"Start Cell Format for :{type}")
            CommonFunction.WriteLog(f"startRow :{startRow}")
            CommonFunction.WriteLog(f"font :{Font}")
            
            TargetWorkSheet = ExcelTargetLib.read_worksheet_as_table(name=type, start=startRow)
            sheetName = ExcelTargetLib.get_active_worksheet()
            CommonFunction.WriteLog(f"sheetName :{sheetName}")
            CommonFunction.WriteLog(f"len(TargetWorkSheet) :{len(TargetWorkSheet)}")
            CommonFunction.WriteLog(f"len(TargetWorkSheet)+1 :{len(TargetWorkSheet)+1}")
            CommonFunction.WriteLog(f"len(TargetWorkSheet)+2 :{len(TargetWorkSheet)+2}")
            
            fontSize = 8
            fontName = Font
            
            if type.lower() == "ag slan":
                
                ExcelTargetLib.set_styles(range_string=F"B{startRow}:B{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"C{startRow}:C{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"D{startRow}:D{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"E{startRow}:E{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=f"F{startRow}:F{len(TargetWorkSheet)}", align_horizontal="right", font_name=fontName, size=fontSize)
                ExcelTargetLib.set_styles(range_string=F"G{startRow}:G{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"H{startRow}:H{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"I{startRow}:I{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"A{startRow}:A{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                
            if type.lower() == "unit slan": 
                    
                ExcelTargetLib.set_styles(range_string=f"E{startRow}:E{len(TargetWorkSheet)}", align_horizontal="right", font_name=fontName, size=fontSize)
                ExcelTargetLib.set_styles(range_string=f"I{startRow}:I{len(TargetWorkSheet)}", align_horizontal="right", font_name=fontName, size=fontSize)
                
                ExcelTargetLib.set_styles(range_string=F"B{startRow}:B{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"C{startRow}:C{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"D{startRow}:D{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"F{startRow}:F{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"G{startRow}:G{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"H{startRow}:H{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"J{startRow}:J{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"K{startRow}:K{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"A{startRow}:A{len(TargetWorkSheet)}", font_name=fontName, size=fontSize )
                
            if type.lower() == "cash flow slan us gaap":
                #Background color: Color [A=255, R=214, G=220, B=228],   Foreground color: Color [WindowText],   Font: [Font: Name=Arial Narrow, Size=8, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: @
                ExcelTargetLib.set_styles(range_string=F"T1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"U1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"V1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"W1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"X1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                #Background color: Color [Window],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=10, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: #,##0;\ \(#,##0\)
                ExcelTargetLib.set_styles(range_string=f"K{startRow}:K{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0;\ \(#,##0\)")
                ExcelTargetLib.set_styles(range_string=f"L{startRow}:L{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0;\ \(#,##0\)")
                #Background color: Color [Window],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=10, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: #,##0.00;\ \(#,##0.00\)
                ExcelTargetLib.set_styles(range_string=f"P{startRow}:P{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"Q{startRow}:Q{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"S{startRow}:S{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"T{startRow}:T{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"U{startRow}:U{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"V{startRow}:V{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"W{startRow}:W{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"X{startRow}:X{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                
                #Background color: Color [Window],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=10, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: @
                ExcelTargetLib.set_styles(range_string=F"B{startRow}:B{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"C{startRow}:C{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"D{startRow}:D{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"E{startRow}:E{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"G{startRow}:G{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"H{startRow}:H{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"I{startRow}:I{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"J{startRow}:J{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"M{startRow}:M{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"N{startRow}:N{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"O{startRow}:O{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"R{startRow}:R{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                #Background color: Color [Window],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=11, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: @
                ExcelTargetLib.set_styles(range_string=F"F{startRow}:F{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=11, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"A{startRow}:A{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
            
            if type.lower() == "cash flow slan ifrs":
                #Background color: Color [A=255, R=214, G=220, B=228],   Foreground color: Color [WindowText],   Font: [Font: Name=Arial Narrow, Size=8, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: @
                ExcelTargetLib.set_styles(range_string=F"B1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"C1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"D1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"E1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"F1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"G1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"H1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"I1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"J1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"K1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"L1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"M1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"N1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"O1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"P1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"Q1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"R1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"T1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"U1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"V1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                #Background color: Color [Window],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=10, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: #,##0.00;\ \(#,##0.00\)
                ExcelTargetLib.set_styles(range_string=f"P{startRow}:P{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"Q{startRow}:Q{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"S{startRow}:S{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"T{startRow}:T{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"U{startRow}:U{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"V{startRow}:V{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                #Background color: Color [Window],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=10, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: #,##0;\ \(#,##0\)
                ExcelTargetLib.set_styles(range_string=f"K{startRow}:K{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0;\ \(#,##0\)")
                ExcelTargetLib.set_styles(range_string=f"L{startRow}:L{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0;\ \(#,##0\)")
                #Background color: Color [Window],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=10, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: @
                ExcelTargetLib.set_styles(range_string=F"B{startRow}:B{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"C{startRow}:C{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"D{startRow}:D{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"E{startRow}:E{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"G{startRow}:G{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"H{startRow}:H{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"I{startRow}:I{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"J{startRow}:J{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"M{startRow}:M{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"N{startRow}:N{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"O{startRow}:O{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"R{startRow}:R{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                #Background color: Color [Window],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=11, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: @
                ExcelTargetLib.set_styles(range_string=F"F{startRow}:F{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=11, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"A{startRow}:A{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
            
            if type.lower() == "liability slan us gaap" or type.lower() == "liability slan ifrs":
                #Background color: Color [A=255, R=214, G=220, B=228],   Foreground color: Color [WindowText],   Font: [Font: Name=Arial Narrow, Size=8, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: @
                ExcelTargetLib.set_styles(range_string=F"B1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"C1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"D1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"E1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"F1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"G1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"H1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"I1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"J1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"K1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"L1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"M1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"N1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"O1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"P1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"R1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"S1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"T1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"U1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"V1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"W1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"X1", cell_fill="FFD700", font_name=Font, size=8, number_format="@")
                #Background color: Color [Window],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=10, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: #,##0.00;\ \(#,##0.00\)
                ExcelTargetLib.set_styles(range_string=f"O{startRow}:O{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"P{startRow}:P{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"R{startRow}:R{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"S{startRow}:S{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"T{startRow}:T{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"U{startRow}:U{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"V{startRow}:V{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"W{startRow}:W{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                ExcelTargetLib.set_styles(range_string=f"X{startRow}:X{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0.00;\ \(#,##0.00\)")
                #Background color: Color [Window],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=10, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: #,##0;\ \(#,##0\)
                ExcelTargetLib.set_styles(range_string=f"K{startRow}:K{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0;\ \(#,##0\)")
                ExcelTargetLib.set_styles(range_string=f"L{startRow}:L{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="#,##0;\ \(#,##0\)")
                #Background color: Color [Window],   Foreground color: Color [WindowText],  Font: [Font: Name=Calibri, Size=10, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: @
                ExcelTargetLib.set_styles(range_string=F"B{startRow}:B{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"C{startRow}:C{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"D{startRow}:D{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"E{startRow}:E{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"G{startRow}:G{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"H{startRow}:H{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"I{startRow}:I{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"J{startRow}:J{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"M{startRow}:M{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"N{startRow}:N{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"Q{startRow}:Q{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
                #Background color: Color [Window],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=11, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: @
                ExcelTargetLib.set_styles(range_string=F"F{startRow}:F{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=11, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"A{startRow}:A{len(TargetWorkSheet)+2}", align_horizontal="right", font_name="Calibri", size=10, number_format="@")
            
            if type.upper() == "FBL3N SAP US GAAP" or type.upper() == "FBL3N SAP IFRS":
                #Background color: Color [A=255, R=128, G=128, B=0],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=11, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code:
                ExcelTargetLib.set_styles(range_string=F"B1", cell_fill="808000", font_name="Calibri", size=11)
                ExcelTargetLib.set_styles(range_string=F"C1", cell_fill="808000", font_name="Calibri", size=11)
                ExcelTargetLib.set_styles(range_string=F"D1", cell_fill="808000", font_name="Calibri", size=11)
                ExcelTargetLib.set_styles(range_string=F"E1", cell_fill="808000", font_name="Calibri", size=11)
                ExcelTargetLib.set_styles(range_string=F"F1", cell_fill="808000", font_name="Calibri", size=11)
                ExcelTargetLib.set_styles(range_string=F"G1", cell_fill="808000", font_name="Calibri", size=11)
                #Background color: Color [A=255, R=255, G=255, B=153],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=11, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: 
                ExcelTargetLib.set_styles(range_string=F"H1", cell_fill="FFFF99", font_name="Calibri", size=11)
                ExcelTargetLib.set_styles(range_string=F"I1", cell_fill="FFFF99", font_name="Calibri", size=11)
                ExcelTargetLib.set_styles(range_string=F"J1", cell_fill="FFFF99", font_name="Calibri", size=11)
                ExcelTargetLib.set_styles(range_string=F"K1", cell_fill="FFFF99", font_name="Calibri", size=11)
                ExcelTargetLib.set_styles(range_string=F"L1", cell_fill="FFFF99", font_name="Calibri", size=11)
                ExcelTargetLib.set_styles(range_string=F"M1", cell_fill="FFFF99", font_name="Calibri", size=11)
                ExcelTargetLib.set_styles(range_string=F"N1", cell_fill="FFFF99", font_name="Calibri", size=11)
                ExcelTargetLib.set_styles(range_string=F"O1", cell_fill="FFFF99", font_name="Calibri", size=11)
                ExcelTargetLib.set_styles(range_string=F"P1", cell_fill="FFFF99", font_name="Calibri", size=11)
                ExcelTargetLib.set_styles(range_string=F"Q1", cell_fill="FFFF99", font_name="Calibri", size=11)
                #Background color: Color [A=255, R=255, G=255, B=153],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=11, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: @
                ExcelTargetLib.set_styles(range_string=F"A{startRow}:A{len(TargetWorkSheet)+1}", cell_fill="FFFF99", font_name="Calibri", size=11, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"B{startRow}:B{len(TargetWorkSheet)+1}", cell_fill="FFFF99", font_name="Calibri", size=11, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"C{startRow}:C{len(TargetWorkSheet)+1}", cell_fill="FFFF99", font_name="Calibri", size=11, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"D{startRow}:D{len(TargetWorkSheet)+1}", cell_fill="FFFF99", font_name="Calibri", size=11, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"E{startRow}:E{len(TargetWorkSheet)+1}", cell_fill="FFFF99", font_name="Calibri", size=11, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"F{startRow}:F{len(TargetWorkSheet)+1}", cell_fill="FFFF99", font_name="Calibri", size=11, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"G{startRow}:G{len(TargetWorkSheet)+1}", cell_fill="FFFF99", font_name="Calibri", size=11, number_format="@")
                #Background color: Color [White],   Foreground color: Color [Black],   Font: [Font: Name=Arial, Size=10, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: 
                #no need to do just different background
                #Background color: Color [Window],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=11, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: #,##0.00
                ExcelTargetLib.set_styles(range_string=f"J{startRow}:J{len(TargetWorkSheet)+1}", font_name="Calibri", size=11, number_format="#,##0.00")
                ExcelTargetLib.set_styles(range_string=f"L{startRow}:L{len(TargetWorkSheet)+1}", font_name="Calibri", size=11, number_format="#,##0.00")
                ExcelTargetLib.set_styles(range_string=f"N{startRow}:N{len(TargetWorkSheet)+1}", font_name="Calibri", size=11, number_format="#,##0.00")
                ExcelTargetLib.set_styles(range_string=f"P{startRow}:P{len(TargetWorkSheet)+1}", font_name="Calibri", size=11, number_format="#,##0.00")
                #Background color: Color [Window],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=11, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: @
                ExcelTargetLib.set_styles(range_string=F"K{startRow}:K{len(TargetWorkSheet)+1}", font_name="Calibri", size=11, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"M{startRow}:M{len(TargetWorkSheet)+1}", font_name="Calibri", size=11, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"O{startRow}:O{len(TargetWorkSheet)+1}", font_name="Calibri", size=11, number_format="@")
                ExcelTargetLib.set_styles(range_string=F"Q{startRow}:Q{len(TargetWorkSheet)+1}", font_name="Calibri", size=11, number_format="@")
                #Background color: Color [Window],   Foreground color: Color [WindowText],   Font: [Font: Name=Calibri, Size=11, Units=3, GdiCharSet=1, GdiVerticalFont=False],   Format code: mm/dd/yyyy
                ExcelTargetLib.set_styles(range_string=F"H{startRow}:H{len(TargetWorkSheet)+1}", font_name="Calibri", size=11, number_format="mm/dd/yyyy")
                ExcelTargetLib.set_styles(range_string=F"I{startRow}:I{len(TargetWorkSheet)+1}", font_name="Calibri", size=11, number_format="mm/dd/yyyy") 
                ExcelTargetLib.set_styles(range_string=F"A1", cell_fill="808000", font_name="Calibri", size=11)
                
            if type.lower() == "depr simulation sap us gaap" or type.lower() == "depr simulation sap ifrs":
                ExcelTargetLib.set_styles(range_string=F"B{startRow}:B{len(TargetWorkSheet)+1}", font_name=fontName, size=fontSize, number_format='#,##0.00', cell_fill="#FFFF99" )
                ExcelTargetLib.set_styles(range_string=F"C{startRow}:C{len(TargetWorkSheet)+1}", font_name=fontName, size=fontSize, number_format='#,##0.00' )
                ExcelTargetLib.set_styles(range_string=F"A{startRow}:A{len(TargetWorkSheet)+1}", font_name=fontName, size=fontSize )
            
            if type.lower() == "asset transaction sap":    
                ExcelTargetLib.set_styles(range_string=F"B{startRow}:B{len(TargetWorkSheet)+1}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"C{startRow}:C{len(TargetWorkSheet)+1}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"D{startRow}:D{len(TargetWorkSheet)+1}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"E{startRow}:E{len(TargetWorkSheet)+1}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"F{startRow}:F{len(TargetWorkSheet)+1}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"G{startRow}:G{len(TargetWorkSheet)+1}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"H{startRow}:H{len(TargetWorkSheet)+1}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"I{startRow}:I{len(TargetWorkSheet)+1}", font_name=fontName, size=fontSize )
                ExcelTargetLib.set_styles(range_string=F"A{startRow}:A{len(TargetWorkSheet)+1}", font_name=fontName, size=fontSize )
            
            CommonFunction.WriteLog(f"Finish Cell Format for :{type}")
        except Exception as errFormat:
            error = str(errFormat)
            CommonFunction.WriteLog(f"Failed set cell formatting error with {error}")
    
    def SetBorder(inWorkBook, inWorkSheet):
        try:
            WorkSheet = inWorkBook[inWorkSheet]
            thin_border = Border(
                left=Side(border_style=BORDER_THIN, color='00000000'),
                right=Side(border_style=BORDER_THIN, color='00000000'),
                top=Side(border_style=BORDER_THIN, color='00000000'),
                bottom=Side(border_style=BORDER_THIN, color='00000000')
            )
            for i, column in enumerate(WorkSheet.columns):
                for j in range(len(column)):
                    Cell = column[j]
                    if (inWorkSheet == "AG SLAN" and (Cell.col_idx == 10 or Cell.col_idx == 11 or Cell.col_idx == 12)) or (inWorkSheet == "Unit SLAN" and Cell.col_idx == 12) or (inWorkSheet == "Completeness check US GAAP" and j < 6) or (inWorkSheet == "Completeness check IFRS" and j < 6):
                        j = j + 1
                        continue
                    else:
                        WorkSheet.cell(row=Cell.row, column=Cell.column).border = thin_border
                        j = j + 1
        except Exception as errBorder:
            error = str(errBorder)
            CommonFunction.WriteLog(f"Failed set border error with {error}")

    # #Main
    def ProcessExcelLeaseRecon(pathAGSLAN, pathUNITSLAN, pathUSGAAP, pathIFRS, pathLiabilityUSGAAP, pathLiabilityIFRS, pathSAPFBL3NUSGAAP, pathSAPFBL3NUSGAAPHTML, pathSAPFBL3NIFRS, pathSAPFBL3NIFRSHTML, pathSAPDeprUSGAAP, pathSAPDeprIFRS, pathSAPAssetTrans, pathSAPAssetTransFix, inCompanyCode, USGAAPFlag, IFRSFlag, reportDate, LastDate, month, year):
        try:
            
            CommonFunction.WriteLog(f"Start copy data to template - {inCompanyCode}")
            CommonFunction.WriteConsole(f"Start copy data to template - {inCompanyCode}")
            ExcelTargetLib.open_workbook(path=LeaseReconTemplatePath)
                        
            CommonFunction.WriteLog(f"pathAGSLAN :{pathAGSLAN}")
            CommonFunction.WriteLog(f"pathAGSLAN: {os.path.exists(str(pathAGSLAN))}")
            
            CommonFunction.WriteLog(f"pathUNITSLAN :{pathUNITSLAN}")
            CommonFunction.WriteLog(f"pathUNITSLAN: {os.path.exists(str(pathUNITSLAN))}")
            
            CommonFunction.WriteLog(f"USGAAPFlag: {USGAAPFlag}")
            CommonFunction.WriteLog(f"pathUSGAAP :{pathUSGAAP}")
            CommonFunction.WriteLog(f"pathUSGAAP :{os.path.exists(str(pathUSGAAP))}")
            CommonFunction.WriteLog(f"pathLiabilityUSGAAP :{pathLiabilityUSGAAP}")
            CommonFunction.WriteLog(f"pathLiabilityUSGAAP: {os.path.exists(str(pathLiabilityUSGAAP))}")
            CommonFunction.WriteLog(f"pathSAPFBL3NUSGAAP :{pathSAPFBL3NUSGAAP}")
            CommonFunction.WriteLog(f"pathSAPFBL3NUSGAAP: {os.path.exists(str(pathSAPFBL3NUSGAAP))}")
            CommonFunction.WriteLog(f"pathSAPDeprUSGAAP :{pathSAPDeprUSGAAP}")
            CommonFunction.WriteLog(f"pathSAPDeprUSGAAP: {os.path.exists(str(pathSAPDeprUSGAAP))}")
            
            CommonFunction.WriteLog(f"IFRSFlag: {IFRSFlag}")
            CommonFunction.WriteLog(f"pathIFRS :{pathIFRS}")
            CommonFunction.WriteLog(f"pathIFRS: {os.path.exists(str(pathIFRS))}")
            CommonFunction.WriteLog(f"pathLiabilityIFRS :{pathLiabilityIFRS}")
            CommonFunction.WriteLog(f"pathLiabilityIFRS: {os.path.exists(str(pathLiabilityIFRS))}")
            CommonFunction.WriteLog(f"pathSAPFBL3NIFRS :{pathSAPFBL3NIFRS}")
            CommonFunction.WriteLog(f"pathSAPFBL3NIFRS: {os.path.exists(str(pathSAPFBL3NIFRS))}")
            CommonFunction.WriteLog(f"pathSAPDeprIFRS :{pathSAPDeprIFRS}")
            CommonFunction.WriteLog(f"pathSAPDeprIFRS: {os.path.exists(str(pathSAPDeprIFRS))}")
            
            CommonFunction.WriteLog(f"pathSAPAssetTrans :{pathSAPAssetTrans}")
            CommonFunction.WriteLog(f"pathSAPAssetTransFix :{pathSAPAssetTransFix}")
            CommonFunction.WriteLog(f"pathSAPAssetTransFix: {os.path.exists(str(pathSAPAssetTransFix))}")
            
            
            # "AG SLAN"
            if os.path.exists(pathAGSLAN):
                CommonFunction.WriteLog(f"Copy Data Active Group")
                Excel.CopyData(pathAGSLAN, Excel.FindStartRowNum(pathAGSLAN,"Activation Group ID","A"), 1, "A2", "", "" , inCompanyCode, "F", month, year)
                Excel.setCellFormat("AG SLAN", 1)
            else:
                Excel.UpdateStatus(inCompanyCode, 'Partially Completed', "Connection issue: Failed Generate Active Group List.")
        
            # "Unit SLAN"
            if os.path.exists(pathUNITSLAN):
                CommonFunction.WriteLog(f"Copy Data Unit List")
                Excel.CopyData(pathUNITSLAN, Excel.FindStartRowNum(pathUNITSLAN,"Unit ID","A"), 2, "A2", "", "", inCompanyCode, "I", month, year)
                Excel.setCellFormat("Unit SLAN", 1)
            else:
                Excel.UpdateStatus(inCompanyCode, 'Partially Completed', "Connection issue: Failed Generate Unit List.")

            if USGAAPFlag == 'x':
                CommonFunction.WriteLog(f"Processing US GAAP")
                # "Cash flow SLAN US GAAP"
                if os.path.exists(pathUSGAAP):
                    CommonFunction.WriteLog(f"Copy Data Cash flow US GAAP")
                    Excel.CopyData(pathUSGAAP, Excel.FindStartRowNum(pathUSGAAP,"Contract ID","A"), 3, "A3", "", "", inCompanyCode, "D", month, year)
                    Excel.setCellFormat("Cash flow SLAN US GAAP", 3)
                else:
                    Excel.UpdateStatus(inCompanyCode, 'Partially Completed', "Connection issue: Failed Generate Cash Flow US GAAP.")
                    
                # "Liability SLAN US GAAP"
                if os.path.exists(pathLiabilityUSGAAP):
                    CommonFunction.WriteLog(f"Copy Data Liability US GAAP")
                    Excel.CopyData(pathLiabilityUSGAAP, Excel.FindStartRowNum(pathLiabilityUSGAAP,"Contract ID","A"), 5, "A3", "", "", inCompanyCode, "D", month, year)
                    Excel.setCellFormat("Liability SLAN US GAAP", 3)
                else:
                    Excel.UpdateStatus(inCompanyCode, 'Partially Completed', "Connection issue: Failed Generate Liability US GAAP.")
                       
                # "FBL3N SAP US GAAP"
                if os.path.exists(pathSAPFBL3NUSGAAP):
                    CommonFunction.WriteLog(f"Copy Data FBL3N US GAAP")
                    Excel.CopyData(pathSAPFBL3NUSGAAP, Excel.FindStartRowNum(pathSAPFBL3NUSGAAP,"Account","B"), 7, "A2", "", "", "", "", month, year)
                    Excel.setCellFormat("FBL3N SAP US GAAP", 2)
                else:
                    Excel.UpdateStatus(inCompanyCode, 'Partially Completed', "Connection issue: Failed Generate FBL3N US GAAP.")
                    
                # Processing Depr Simulation SAP US GAAP
                if os.path.exists(pathSAPDeprUSGAAP):
                    CommonFunction.WriteLog(f"Copy Data Depr US GAAP")
                    Excel.CopyData(pathSAPDeprUSGAAP, 1, 9, "B2", "", "Object", "", "", month, year)
                    TargetWorkSheet = ExcelTargetLib.read_worksheet_as_table(name="Depr Simulation SAP US GAAP", start=2)
                    ExcelTargetLib.set_cell_formula(range_string=f"A2:A{len(TargetWorkSheet)+1}", formula=f"=VALUE(LEFT(B2,{assetNbrLength}))", transpose= True)
                    Excel.setCellFormat("Depr Simulation SAP US GAAP", 2)
                else:
                    Excel.UpdateStatus(inCompanyCode, 'Partially Completed', "Connection issue: Failed Generate Depreciation Simulation US GAAP.")
                    
            if IFRSFlag == 'x':
                CommonFunction.WriteLog(f"Processing IFRS")
                # "Cash flow SLAN IFRS"
                if os.path.exists(pathIFRS):
                    CommonFunction.WriteLog(f"Copy Data Cash flow IFRS")
                    Excel.CopyData(pathIFRS, Excel.FindStartRowNum(pathIFRS,"Contract ID","A"), 4, "A3", "", "", inCompanyCode, "D", month, year)
                    Excel.setCellFormat("Cash flow SLAN IFRS", 3)
                else:
                    Excel.UpdateStatus(inCompanyCode, 'Partially Completed', "Connection issue: Failed Generate Cash Flow IFRS.")
                    
                # "Liability SLAN IFRS"
                if os.path.exists(pathLiabilityIFRS):
                    CommonFunction.WriteLog(f"Copy Data Liability IFRS")
                    Excel.CopyData(pathLiabilityIFRS, Excel.FindStartRowNum(pathLiabilityIFRS,"Contract ID","A"), 6, "A3", "", "", inCompanyCode, "D", month, year)
                    Excel.setCellFormat("Liability SLAN IFRS", 3)
                else:
                    Excel.UpdateStatus(inCompanyCode, 'Partially Completed', "Connection issue: Failed Generate Liability IFRS.")
                    
                # "FBL3N SAP US IFRS"
                if os.path.exists(pathSAPFBL3NIFRS):
                    CommonFunction.WriteLog(f"Copy Data FBL3N IFRS")
                    Excel.CopyData(pathSAPFBL3NIFRS, Excel.FindStartRowNum(pathSAPFBL3NIFRS,"Account","B"), 8, "A2", "", "", "", "", month, year)
                    Excel.setCellFormat("FBL3N SAP IFRS", 2)
                else:
                    Excel.UpdateStatus(inCompanyCode, 'Partially Completed', "Connection issue: Failed Generate FBL3N IFRS.")

                # Processing Depr Simulation SAP IFRS
                if os.path.exists(pathSAPDeprIFRS):
                    CommonFunction.WriteLog(f"Copy Data Depr IFRS")
                    Excel.CopyData(pathSAPDeprIFRS, 1, 10, "B2", "", "Object", "", "", month, year)
                    TargetWorkSheet = ExcelTargetLib.read_worksheet_as_table(name="Depr Simulation SAP IFRS", start=2)
                    ExcelTargetLib.set_cell_formula(range_string=f"A2:A{len(TargetWorkSheet)+1}", formula=f"=VALUE(LEFT(B2,{assetNbrLength}))", transpose=True)
                    Excel.setCellFormat("Depr Simulation SAP IFRS", 2)
                else:
                    Excel.UpdateStatus(inCompanyCode, 'Partially Completed', "Connection issue: Failed Generate Depreciation Simulation IFRS.")

            # Processing Asset Transaction SAP
            if os.path.exists(pathSAPAssetTrans):
                CommonFunction.WriteLog(f"Copy Data Asset Transaction SAP")
                Excel.convertToXLSX(pathSAPAssetTrans, pathSAPAssetTransFix)
                Excel.CopyData(pathSAPAssetTransFix, 1, 11, "A1", "", "", "", "", month, year)
                Excel.ProcessingAssetTransactionSAP(month)
                Excel.setCellFormat("Asset Transaction SAP", 2)
            else:
                Excel.UpdateStatus(inCompanyCode, 'Asset', 'Asset Transaction empty')

            if USGAAPFlag == 'x':
                # Processing Completeness check US GAAP
                try:
                    Excel.ProcessingCompletenesscheckUSGAAP(inCompanyCode, LastDate)
                except Exception as error:
                    Excel.UpdateStatus(inCompanyCode, 'Partially Completed', "Other Reason: Failed Completeness check US GAAP.")
                    CommonFunction.WriteLog(f"Processing Completeness check US GAAP error with: {error}")

            if IFRSFlag == 'x':
                # Processing Completeness check IFRS
                try:
                    Excel.ProcessingCompletenesscheckIFRS(inCompanyCode, LastDate)
                except Exception as error:
                    Excel.UpdateStatus(inCompanyCode, 'Partially Completed', "Other Reason: Failed Completeness check IFRS.")
                    CommonFunction.WriteLog(f"Processing Completeness check IFRS error with: {error}")
            
                
            TargetWorkSheet = ExcelTargetLib.read_worksheet_as_table(name="Completeness check US GAAP", start=1)
            ContractIDList = TablesTargetLib.get_table_column(table=TargetWorkSheet, column="A")
            
            # Save reconciliation according to Naming convention: YYYY-MM CCCC Lease
            # Completeness check **** YYYY- year; MM- month; CCCC- company code; **** –date and
            # hour of finish robot work
            # prefixName = datetime.datetime.now().strftime("%Y-%m")
            # dateName = datetime.datetime.now().strftime("%Y-%m-%d")
            
            prefixName = reportDate.format("%Y-%m")
            dateName = reportDate.format("%Y-%m-%d")
            hourName = datetime.datetime.now().strftime("%H_%M")
            
            ResultFileName = f'{prefixName} {inCompanyCode} Lease Completeness check {dateName} {hourName}.xlsx'
            ResultFile = pathResultSummary+"\\"+ResultFileName
            
            ExcelTargetLib.save_workbook(path=ResultFile)
            ExcelTargetLib.close_workbook()
            # Lock File
            CommonFunction.WriteLog(f"Before Load {str(ResultFile)}")
            WorkBook = openpyxl.load_workbook(filename=str(ResultFile))
            CommonFunction.WriteLog(f"After Load")
            
            Excel.SetBorder(WorkBook,"AG SLAN")
            Excel.SetBorder(WorkBook,"Unit SLAN")

            Excel.SetBorder(WorkBook,"Cash flow SLAN US GAAP")
            Excel.SetBorder(WorkBook,"Liability SLAN US GAAP")
            Excel.SetBorder(WorkBook,"FBL3N SAP US GAAP")
            Excel.SetBorder(WorkBook,"Depr Simulation SAP US GAAP")
            Excel.SetBorder(WorkBook,"Completeness check US GAAP")

            Excel.SetBorder(WorkBook,"Cash flow SLAN IFRS")
            Excel.SetBorder(WorkBook,"Liability SLAN IFRS")
            Excel.SetBorder(WorkBook,"FBL3N SAP IFRS")
            Excel.SetBorder(WorkBook,"Depr Simulation SAP IFRS")
            Excel.SetBorder(WorkBook,"Completeness check IFRS")
            
            if  USGAAPFlag != 'x':
                Excel.HideSheet(WorkBook,"Cash flow SLAN US GAAP")
                Excel.HideSheet(WorkBook,"Liability SLAN US GAAP")
                Excel.HideSheet(WorkBook,"FBL3N SAP US GAAP")
                Excel.HideSheet(WorkBook,"Depr Simulation SAP US GAAP")
                Excel.HideSheet(WorkBook,"Completeness check US GAAP")
            
            if IFRSFlag != 'x':
                Excel.HideSheet(WorkBook,"Cash flow SLAN IFRS")
                Excel.HideSheet(WorkBook,"Liability SLAN IFRS")
                Excel.HideSheet(WorkBook,"FBL3N SAP IFRS")
                Excel.HideSheet(WorkBook,"Depr Simulation SAP IFRS")
                Excel.HideSheet(WorkBook,"Completeness check IFRS")
                
            Excel.ProtectFiles(WorkBook, FilePassword, ContractIDList)
            WorkBook.save(filename=str(ResultFile))
            WorkBook.close()
            
            
            ExcelTargetLib.open_workbook(path=ResultFile)
            
            
            act = ExcelTargetLib.get_active_worksheet()
            CommonFunction.WriteLog(f"active sheet before: {act}")
            if IFRSFlag == 'x':
                ExcelTargetLib.set_active_worksheet("Completeness check IFRS")
            else:
                ExcelTargetLib.set_active_worksheet("Completeness check US GAAP")
                
            act2 = ExcelTargetLib.get_active_worksheet()
            CommonFunction.WriteLog(f"active sheet after: {act2}")
            
            ExcelTargetLib.save_workbook(path=ResultFile)
            ExcelTargetLib.close_workbook()
            
            CommonFunction.WriteLog(f"Finish copy data to template.")
            CommonFunction.WriteConsole(f"Finish copy data to template.")
            
            CommonFunction.WriteConsole(f"Start Upload to Sharepoint.")
            UploadFile(fileName=ResultFileName, filePath=ResultFile, typeFile="summary")
            
            if sharepointFolderDetail != None and sharepointFolderDetail !='None':
                UploadFile(fileName=pathAGSLAN.replace(pathResultExcel, ''), filePath=pathAGSLAN, typeFile="detail")
                UploadFile(fileName=pathUNITSLAN.replace(pathResultExcel, ''), filePath=pathUNITSLAN, typeFile="detail")
                
                UploadFile(fileName=pathUSGAAP.replace(pathResultExcel, ''), filePath=pathUSGAAP, typeFile="detail")
                UploadFile(fileName=pathLiabilityUSGAAP.replace(pathResultExcel, ''), filePath=pathLiabilityUSGAAP, typeFile="detail")
                UploadFile(fileName=pathSAPFBL3NUSGAAP.replace(pathResultSAP, ''), filePath=pathSAPFBL3NUSGAAP, typeFile="detail")
                UploadFile(fileName=pathSAPDeprUSGAAP.replace(pathResultSAP, ''), filePath=pathSAPDeprUSGAAP, typeFile="detail")
                
                if IFRSFlag == 'x':
                    UploadFile(fileName=pathIFRS.replace(pathResultExcel, ''), filePath=pathIFRS, typeFile="detail")
                    UploadFile(fileName=pathLiabilityIFRS.replace(pathResultExcel, ''), filePath=pathLiabilityIFRS, typeFile="detail")
                    UploadFile(fileName=pathSAPFBL3NIFRS.replace(pathResultSAP, ''), filePath=pathSAPFBL3NIFRS, typeFile="detail")
                    UploadFile(fileName=pathSAPDeprIFRS.replace(pathResultSAP, ''), filePath=pathSAPDeprIFRS, typeFile="detail")
                
                UploadFile(fileName=pathSAPAssetTransFix.replace(pathResultSAP, ''), filePath=pathSAPAssetTransFix, typeFile="detail")
            CommonFunction.WriteConsole(f"Finish Upload to Sharepoint.")


            if (USGAAPFlag == 'x' and IFRSFlag == 'x' ):
                if (os.path.exists(str(pathAGSLAN)) and os.path.exists(str(pathUNITSLAN)) and os.path.exists(str(pathUSGAAP)) and os.path.exists(str(pathLiabilityUSGAAP)) and os.path.exists(str(pathSAPFBL3NUSGAAP)) and os.path.exists(str(pathSAPFBL3NUSGAAPHTML)) and os.path.exists(str(pathSAPDeprUSGAAP)) and os.path.exists(str(pathIFRS)) and os.path.exists(str(pathLiabilityIFRS)) and os.path.exists(str(pathSAPFBL3NIFRS)) and os.path.exists(str(pathSAPFBL3NIFRSHTML)) and os.path.exists(str(pathSAPDeprIFRS))  ) :
                    if not os.path.exists(pathSAPAssetTransFix):
                        Excel.UpdateStatus(inCompanyCode, 'Completed', 'All file downloaded. Asset Transaction empty')
                    else:
                        Excel.UpdateStatus(inCompanyCode, 'Completed', 'All file downloaded.')
                if not os.path.exists(str(pathAGSLAN)) and not os.path.exists(str(pathUNITSLAN)) and not os.path.exists(str(pathUSGAAP)) and not os.path.exists(str(pathLiabilityUSGAAP)) and not os.path.exists(str(pathSAPFBL3NUSGAAP)) and not os.path.exists(str(pathSAPFBL3NUSGAAPHTML)) and not os.path.exists(str(pathSAPDeprUSGAAP)) and not os.path.exists(str(pathIFRS)) and not os.path.exists(str(pathLiabilityIFRS)) and not os.path.exists(str(pathSAPFBL3NIFRS)) and not os.path.exists(str(pathSAPFBL3NIFRSHTML))  and not os.path.exists(str(pathSAPDeprIFRS)):
                    Excel.UpdateStatus(inCompanyCode, 'Incomplete', 'All file not downloaded.')

            if (USGAAPFlag == 'x' and IFRSFlag != 'x' ):
                if (os.path.exists(str(pathAGSLAN)) and os.path.exists(str(pathUNITSLAN)) and os.path.exists(str(pathUSGAAP)) and os.path.exists(str(pathLiabilityUSGAAP)) and os.path.exists(str(pathSAPFBL3NUSGAAP)) and os.path.exists(str(pathSAPFBL3NUSGAAPHTML)) and os.path.exists(str(pathSAPDeprUSGAAP))):
                    if not os.path.exists(pathSAPAssetTransFix):
                        Excel.UpdateStatus(inCompanyCode, 'Completed', 'All file downloaded. Asset Transaction empty')
                    else:
                        Excel.UpdateStatus(inCompanyCode, 'Completed', 'All file downloaded.')
                if not os.path.exists(str(pathAGSLAN)) and not os.path.exists(str(pathUNITSLAN)) and not os.path.exists(str(pathUSGAAP)) and not os.path.exists(str(pathLiabilityUSGAAP)) and not os.path.exists(str(pathSAPFBL3NUSGAAP)) and not os.path.exists(str(pathSAPFBL3NUSGAAPHTML)) and not os.path.exists(str(pathSAPDeprUSGAAP)):
                    Excel.UpdateStatus(inCompanyCode, 'Incomplete', 'All file not downloaded.')

            if (USGAAPFlag != 'x' and IFRSFlag == 'x' ):
                if (os.path.exists(str(pathAGSLAN)) and os.path.exists(str(pathUNITSLAN)) and os.path.exists(str(pathIFRS)) and os.path.exists(str(pathLiabilityIFRS)) and os.path.exists(str(pathSAPFBL3NIFRS)) and os.path.exists(str(pathSAPFBL3NIFRSHTML)) and os.path.exists(str(pathSAPDeprIFRS))  ) :
                    if not os.path.exists(pathSAPAssetTransFix):
                        Excel.UpdateStatus(inCompanyCode, 'Completed', 'All file downloaded. Asset Transaction empty')
                    else:
                        Excel.UpdateStatus(inCompanyCode, 'Completed', 'All file downloaded.')
                if not os.path.exists(str(pathAGSLAN)) and not os.path.exists(str(pathUNITSLAN)) and not os.path.exists(str(pathIFRS)) and not os.path.exists(str(pathLiabilityIFRS)) and not os.path.exists(str(pathSAPFBL3NIFRS)) and not os.path.exists(str(pathSAPFBL3NIFRSHTML)) and not os.path.exists(str(pathSAPDeprIFRS)):
                    Excel.UpdateStatus(inCompanyCode, 'Incomplete', 'All file not downloaded.')
                
            return excelPath.pathResultFile
        except Exception as error:
            Excel.UpdateStatus(inCompanyCode, 'Partially Complete', 'Other Reason: Process copy data to Excel Result error')
            CommonFunction.WriteLog(f"Process copy data to Excel Result error with: {error}")
    
    def UpdateStatus(companyCode, statusReport, errorMessage):
        try:
            CommonFunction.WriteLog(f"Update Status Company: {str(companyCode)}")
            ExcelLib.open_workbook(path=LeaseReconSchedulerPath)
            rows = ExcelLib.read_worksheet('Sheet1', start=1)
            i = 1
            for cell in rows:
                code = cell['A'] #company code - 1002
                
                errorMsg    =   ''
                status      =   ''
                
                if cell.get('D') != None:
                    status = cell['D']

                if cell.get('E') != None:
                    errorMsg = cell['E']

                if i == 1:
                    ExcelLib.set_cell_value(row=i, column="D", value="Status")
                    ExcelLib.set_cell_value(row=i, column="E", value="Error Message")

                if str(code) == companyCode:
                    CommonFunction.WriteLog(f"New Status: {str(statusReport)}")
                    CommonFunction.WriteLog(f"New Message: {str(errorMessage)}")
                    CommonFunction.WriteLog(f"Current Status: {str(status)}")
                    CommonFunction.WriteLog(f"Current Message: {str(errorMsg)}")
                    
                    if statusReport == "Completed" or statusReport == "Incomplete":
                        ExcelLib.set_cell_value(row=i, column="D", value=statusReport)
                        ExcelLib.set_cell_value(row=i, column="E", value=errorMessage)
                    else:
                        if statusReport != "Asset":
                            ExcelLib.set_cell_value(row=i, column="D", value=statusReport)
                        ExcelLib.set_cell_value(row=i, column="E", value=str(errorMsg)+" "+errorMessage)
                i=i+1
            ExcelLib.save_workbook(path=LeaseReconSchedulerPath)
            CommonFunction.WriteLog(f"Finish Update Status Company: {companyCode}")
        except Exception as errStatus:
            error = str(errStatus)
            CommonFunction.Writelog(f"Failed to update status error with {error}")

    
